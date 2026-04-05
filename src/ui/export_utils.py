"""Shared export utilities for Excel and PDF"""

import os
from datetime import datetime
from pathlib import Path
from tkinter import messagebox
from typing import Dict, List, Optional

try:
    from matplotlib.backends.backend_pdf import PdfPages
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


def get_export_folder(source_file_path: Optional[str] = None) -> Path:
    """Get or create date-stamped export folder next to source data files
    Reuses folder if already exists for that day
    
    Args:
        source_file_path: Path to source file (CBOM, YMBD, etc.) - uses its parent
        
    Returns:
        Path to export folder with date stamp: mm-dd-yyyy
    """
    # Create date-only folder name
    date_stamp = datetime.now().strftime("%m-%d-%Y")
    folder_name = f"Exports_{date_stamp}"
    
    if source_file_path and os.path.exists(source_file_path):
        # Create export folder next to source file
        parent_dir = Path(source_file_path).parent
        export_folder = parent_dir / folder_name
    else:
        # Fallback to current working directory
        export_folder = Path.cwd() / folder_name
    
    # Create folder if it doesn't exist (reuse if already exists)
    export_folder.mkdir(parents=True, exist_ok=True)
    
    return export_folder


def export_screen_to_pdf(
    widget,
    export_folder: Path,
    filename_prefix: str = "screenshot",
    title: Optional[str] = None
) -> bool:
    """Export screenshot of widget/screen to PDF
    
    Args:
        widget: The tkinter widget to capture
        export_folder: Path to export folder
        filename_prefix: Prefix for filename
        title: Optional title for the PDF
        
    Returns:
        True if export succeeded, False otherwise
    """
    try:
        from PIL import ImageGrab
        from matplotlib.backends.backend_pdf import PdfPages
        import matplotlib.pyplot as plt
    except ImportError:
        messagebox.showerror("Export Failed", "PIL or Matplotlib is required for screenshot export")
        return False
    
    timestamp = datetime.now().strftime("%H-%M-%S")
    filename = f"{filename_prefix}_{timestamp}.pdf"
    file_path = export_folder / filename
    
    try:
        # Get widget position and size
        x = widget.winfo_rootx()
        y = widget.winfo_rooty()
        width = widget.winfo_width()
        height = widget.winfo_height()
        
        # Capture screenshot
        screenshot = ImageGrab.grab(bbox=(x, y, x + width, y + height))
        
        # Save to PDF
        with PdfPages(file_path) as pdf:
            fig = plt.figure(figsize=(screenshot.width / 100, screenshot.height / 100), dpi=100)
            ax = fig.add_subplot(111)
            ax.imshow(screenshot)
            ax.axis('off')
            
            if title:
                fig.suptitle(title, fontsize=16, fontweight='bold', y=0.98)
            
            plt.tight_layout()
            pdf.savefig(fig, bbox_inches='tight')
            
            # Add metadata
            d = pdf.infodict()
            d['Title'] = title or 'Screen Capture'
            d['Author'] = 'Performance Center'
            d['Creator'] = 'Performance Center Application'
            d['CreationDate'] = datetime.now()
        
        plt.close(fig)
        messagebox.showinfo("Export Successful", f"Screenshot exported to:\n{file_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Export Failed", f"Error capturing screenshot:\n{str(e)}")
        return False


def export_chart_to_pdf(
    figure,
    export_folder: Path,
    filename_prefix: str = "chart",
    title: Optional[str] = None
) -> bool:
    """Export matplotlib figure to PDF in predetermined folder
    
    Args:
        figure: Matplotlib figure object
        export_folder: Path to export folder
        filename_prefix: Prefix for filename
        title: Optional title for the PDF
        
    Returns:
        True if export succeeded, False otherwise
    """
    if not HAS_MATPLOTLIB:
        messagebox.showerror("Export Failed", "Matplotlib is not available for PDF export")
        return False
    
    timestamp = datetime.now().strftime("%H-%M-%S")
    filename = f"{filename_prefix}_{timestamp}.pdf"
    file_path = export_folder / filename
    
    try:
        with PdfPages(file_path) as pdf:
            # Save the figure to PDF
            pdf.savefig(figure, bbox_inches='tight')
            
            # Add metadata
            d = pdf.infodict()
            d['Title'] = title or 'Performance Chart'
            d['Author'] = 'Performance Center'
            d['Creator'] = 'Performance Center Application'
            d['CreationDate'] = datetime.now()
        
        messagebox.showinfo("Export Successful", f"Chart exported to:\n{file_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Export Failed", f"Error exporting chart:\n{str(e)}")
        return False


def export_data_to_excel(
    data: Dict[int, Dict[str, int]],
    periods: List[str],
    years: List[int],
    export_folder: Path,
    filename_prefix: str,
    entity_count: int = 0,
    mode: str = "",
    granularity: str = "",
    selected_entity_ids: Optional[List[str]] = None
) -> bool:
    """Export data to Excel file in predetermined folder
    
    Args:
        data: Dictionary {year: {period: value}}
        periods: List of period labels
        years: List of selected years
        export_folder: Path to export folder
        filename_prefix: Prefix for filename
        entity_count: Number of entities selected
        mode: Analysis mode ("12nc" or "room")
        granularity: Time granularity
        
    Returns:
        True if export succeeded, False otherwise
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messagebox.showerror(
            "Export Failed", 
            "openpyxl library is required for Excel export.\nPlease install it with: pip install openpyxl"
        )
        return False
    
    timestamp = datetime.now().strftime("%H-%M-%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = export_folder / filename
    
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if ws is None:
            raise Exception("Failed to create worksheet")
        
        ws.title = "Bulk Analysis"
        
        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="4A8F93", end_color="4A8F93", fill_type="solid")
        
        # Title section
        ws['A1'] = 'Analysis Type'
        ws['B1'] = f'Bulk {mode.upper()} Analysis'
        ws['A1'].font = title_font
        
        ws['A2'] = 'Selected Entities'
        ws['B2'] = entity_count
        
        ws['A3'] = 'Granularity'
        ws['B3'] = granularity
        
        ws['A4'] = 'Export Date'
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Add selected entity IDs if provided
        current_row = 5
        if selected_entity_ids:
            ws.cell(current_row, 1, 'Selected Entities').font = header_font
            # List entity IDs
            for idx, entity_id in enumerate(selected_entity_ids, start=1):
                ws.cell(current_row, idx + 1, entity_id)
            current_row += 1
        
        current_row += 1  # Blank row
        
        # Data table header
        start_row = current_row
        ws.cell(start_row, 1, 'Period').font = header_font
        ws.cell(start_row, 1).fill = header_fill
        
        for col_idx, year in enumerate(sorted(years), start=2):
            cell = ws.cell(start_row, col_idx, str(year))
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        for row_idx, period in enumerate(periods, start=start_row + 1):
            ws.cell(row_idx, 1, period)
            for col_idx, year in enumerate(sorted(years), start=2):
                value = data.get(year, {}).get(period, 0)
                ws.cell(row_idx, col_idx, int(value))
        
        # Summary section
        summary_row = start_row + len(periods) + 2
        ws.cell(summary_row, 1, 'Total').font = header_font
        total = sum(data.get(year, {}).get(period, 0) 
                   for year in years 
                   for period in periods)
        ws.cell(summary_row, 2, int(total))
        
        ws.cell(summary_row + 1, 1, 'Average per Entity').font = header_font
        ws.cell(summary_row + 1, 2, int(total / entity_count) if entity_count else 0)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
        
        # Save workbook
        wb.save(file_path)
        
        messagebox.showinfo("Export Successful", f"Data exported to:\n{file_path}")
        return True
        
    except Exception as e:
        messagebox.showerror("Export Failed", f"Error exporting to Excel:\n{str(e)}")
        return False
