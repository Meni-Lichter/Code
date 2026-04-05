"""Details Panel - Shows entity metadata and information"""

import customtkinter as ctk
from tkinter import messagebox
from pathlib import Path
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DetailsPanel:
    """Manages the Details panel content and updates"""
    
    def __init__(self, panel_widget, colors, font_sizes, get_font_func):
        """Initialize the details panel manager
        Args:
            panel_widget: The CTkFrame panel widget
            colors: Dictionary of color constants
            font_sizes: Dictionary of font sizes
            get_font_func: Function to get cached fonts
        Does: Sets up the details panel with necessary references and styling information
        Returns: None
        """
        self.panel = panel_widget
        self.COLORS = colors
        self.FONT_SIZES = font_sizes
        self._get_font = get_font_func
        self.content_frame = None
    
    def update(self, entity_obj, mode):
        """Update panel with entity information
        Args:
            entity_obj: Room or TwelveNC object
            mode: Current mode ('room' or '12nc')
        Does: Updates the details panel with information from the given entity object
        Returns: None
        """
        # Find the content frame in the panel
        self.content_frame = self._find_content_frame()
        
        if not self.content_frame:
            return
        
        # Clear existing content
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # Create details container
        details_container = ctk.CTkFrame(
            self.content_frame,
            fg_color="transparent"
        )
        details_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Entity ID and Description
        self._add_detail_row(details_container, "ID:", entity_obj.id, is_header=True)
        self._add_detail_row(details_container, "Description:", entity_obj.description)
        
        # Mode-specific details
        if mode == "room":
            self._add_room_details(details_container, entity_obj)
        elif mode == "12nc":  # 12nc mode
            self._add_12nc_details(details_container, entity_obj)
        
        # Sales history summary
        self._add_detail_separator(details_container)
        sales_count = len(entity_obj.sales_history)
        total_quantity = sum(record.quantity for record in entity_obj.sales_history) if entity_obj.sales_history else 0
        self._add_detail_row(details_container, "Sales Records:", str(sales_count))
        self._add_detail_row(details_container, "Total Sold Quantity:", str(total_quantity))
    
    def _find_content_frame(self):
        """Find the content frame within the panel widget
        Args: None
        Does: Searches through the panel's child widgets to find the content frame where details should be displayed
        Returns: The content frame widget if found, otherwise None
        """
        # Look for the content frame at grid row=1
        for child in self.panel.winfo_children():
            if isinstance(child, ctk.CTkFrame):
                try:
                    grid_info = child.grid_info()
                    if grid_info and grid_info.get('row') == 1:
                        return child
                except:
                    continue
        return None
    
    def _add_detail_row(self, parent, label, value, is_header=False):
        """Add a label-value row to the details panel
        Args:
            parent: Parent widget
            label: Label text
            value: Value text
            is_header: Whether this is a header row (larger font)
        Does: Creates a row in the details panel with the given label and value, styled according to whether it's a header or not
        Returns: None
        """
        row_frame = ctk.CTkFrame(parent, fg_color="transparent")
        row_frame.pack(fill="x", pady=5)
        
        font_size = self.FONT_SIZES["title"] if is_header else self.FONT_SIZES["body"]
        font_weight = "bold" if is_header else "bold"
        
        # Label
        label_widget = ctk.CTkLabel(
            row_frame,
            text=label,
            font=self._get_font(size=font_size, weight=font_weight),
            text_color=self.COLORS["text_dark"],
            anchor="w"
        )
        label_widget.pack(side="left", padx=(0, 10))
        
        # Value (use Entry for copyable text)
        value_font_weight = "bold" if is_header else "normal"
        value_widget = ctk.CTkEntry(
            row_frame,
            font=self._get_font(size=font_size, weight=value_font_weight),
            text_color=self.COLORS["text_muted"] if not is_header else self.COLORS["accent_teal"],
            fg_color="transparent",
            border_width=0,
            state="normal"  # Allow selection and copying
        )
        value_widget.insert(0, value)
        value_widget.configure(state="readonly")  # Prevent editing but allow copying
        value_widget.pack(side="left", fill="x", expand=True)
    
    def _add_detail_separator(self, parent):
        """Add a visual separator line
        Args:
            parent: Parent widget
        Does: Adds a horizontal line to separate sections in the details panel
        Returns: None
        """
        separator = ctk.CTkFrame(parent, height=1, fg_color=self.COLORS["border"])
        separator.pack(fill="x", pady=15)
    
    def _add_room_details(self, parent, room_obj):
        """Add room-specific details
        Args:
            parent: Parent widget
            room_obj: Room object
        Does: Adds details specific to a room entity, such as number of components and total items
        Returns: None
        """
        self._add_detail_separator(parent)
        self._add_detail_row(parent, "Total Components:", str(len(room_obj.components)))
        self._add_detail_row(parent, "Overall Component units:", str(room_obj.total_items))
    
    def _add_12nc_details(self, parent, nc12_obj):
        """Add 12NC-specific details
        Args:
            parent: Parent widget
            nc12_obj: TwelveNC object
        Does: Adds details specific to a 12NC entity, such as IGT, number of rooms deployed in, and total quantity
        Returns: None
        """
        self._add_detail_separator(parent)
        self._add_detail_row(parent, "IGT #:", nc12_obj.igt if nc12_obj.igt else "N/A")
        self._add_detail_row(parent, "# of Rooms Deployed in :", str(len(nc12_obj.components)))
        self._add_detail_row(parent, "Overall Quantity:", str(nc12_obj.total_items))
    
    def export_to_excel(self, entity, export_folder, mode):
        """Export details panel data to Excel
        
        Args:
            entity: The entity object to export
            export_folder: Path to export folder
            mode: Current mode ("12nc" or "room")
        """
        if not HAS_OPENPYXL:
            messagebox.showerror("Export Failed", "openpyxl is required for Excel export")
            return
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            messagebox.showerror("Export Failed", "Failed to create Excel worksheet")
            return
        ws.title = "Details"
        
        # Styles
        header_fill = PatternFill(start_color="4A8F93", end_color="4A8F93", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        label_font = Font(bold=True, size=11)
        
        # Title
        entity_id = entity.id
        safe_entity_id = entity_id.replace('/', '_').replace('\\', '_')
        mode_text = "12NC" if mode == "12nc" else "Room"
        ws['A1'] = f"{mode_text} Details - {entity_id}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Basic Info
        row = 3
        ws[f'A{row}'] = "Basic Information"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "ID:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = entity.id
        
        row += 1
        ws[f'A{row}'] = "Description:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = entity.description
        
        # Mode-specific details
        row += 2
        ws[f'A{row}'] = f"{mode_text} Specific Details"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        if mode == "room":
            ws[f'A{row}'] = "Total Components:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = len(entity.components)
            
            row += 1
            ws[f'A{row}'] = "Overall Component units:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = entity.total_items
        else:  # 12nc
            ws[f'A{row}'] = "IGT #:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = entity.igt if entity.igt else "N/A"
            
            row += 1
            ws[f'A{row}'] = "# of Rooms Deployed in:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = len(entity.components)
            
            row += 1
            ws[f'A{row}'] = "Overall Quantity:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = entity.total_items
        
        # Sales Summary
        row += 2
        ws[f'A{row}'] = "Sales Summary"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        sales_count = len(entity.sales_history)
        ws[f'A{row}'] = "Sales Records:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = sales_count
        
        row += 1
        total_quantity = sum(record.quantity for record in entity.sales_history) if entity.sales_history else 0
        ws[f'A{row}'] = "Total Sold Quantity:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = total_quantity
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        
        # Save file
        timestamp = datetime.now().strftime("%H-%M-%S")
        filename = f"details_{safe_entity_id}_{timestamp}.xlsx"
        file_path = export_folder / filename
        
        try:
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Details exported to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Error saving file:\n{str(e)}")
