
from logging import config
import pandas as pd
import re
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
import logging
from tkinter import messagebox
import json


def read_file(path: Path, file_type:str ,header = None) -> pd.DataFrame:
    """
    Reads an Excel or CSV file into a DataFrame, handling different formats and errors.
    
    Args: 
        path (Path): The path to the file to read.
        file_type (str): The type of the file ('excel' or 'csv').
        header (int, list of int, None): Row(s) to use as the column names. Defaults to None.
    Returns:
        pd.DataFrame: The contents of the file as a DataFrame, or None if an error occurred.
    """

    try:
        if not os.path.exists(path):
            messagebox.showerror(
                "File Not Found",
                f"The specified file does not exist:\n{path}\n\nPlease check the file path and try again."
            )
            return None
        if file_in_use(path):
            messagebox.showerror(
                "File In Use",
                f"The specified file is currently open in another program:\n{path}\n\nPlease close the file and try again."
            )
            return None
        
        if isinstance(path, str):
            path = Path(path)
        ext = path.suffix.lower()
    
        relevant_sheet = pick_sheet(path, file_type)
        print (f"Using sheet: {relevant_sheet} from file: {path.name}")

        # Read file based on extension
        if ext == ".csv":
            df = pd.read_csv(path, header=header)
        elif ext in [".xlsx", ".xlsm"]: 
            df = pd.read_excel(path, sheet_name=relevant_sheet, header=header, engine='openpyxl')
        elif ext in [".xls"]:
            df = pd.read_excel(path, sheet_name=relevant_sheet, header=header,engine='xlrd')
            print ("read .xls file with pandas")
        else:
            raise ValueError(f"Unsupported file format: {ext}. Only .xlsx, .xlsm, and .csv files are supported.")

        print(f"DataFrame shape: {df.shape}")
        
        required_columns = config[file_type].get("required_fields", [])

        if not set(required_columns).issubset(set(df.columns)):
            messagebox.showerror(
                "Error",
                f"Sheet '{relevant_sheet}' must contain columns: {required_columns}"
            )
            return None

        return df
    
    except Exception as e:
        messagebox.showerror(
            "Error",
            f"Could not read file:\n{e}"
        )
        return None




def col_letter_to_index(col_letter):
    """Converts Excel column letter(s) to a zero-based index.
    Args:
        col_letter (str): Excel column letter(s) (e.g., 'A', 'B', ..., 'Z', 'AA', etc.)
    
    Returns:
        int: Zero-based column index
    """

    col_letter = col_letter.upper()
    index = 0
    for i, char in enumerate(reversed(col_letter)):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1


def normalize_identifier(value):
    """
    Normalize an identifier (like room number or 12NC) by:
    - Converting to string
    - Removing spaces, hyphens, and underscores
    - Stripping leading and trailing whitespace
    input : 
        - value: The value to normalize (can be any type, will be converted to string)
    output:
        - string: The normalized string with non-alphanumeric characters removed.
    """
    if pd.isna(value):
        return ""
    
    # Convert to string and remove all non-alphanumeric characters
    normalized = str(value.replace(" ", "").replace("-", "").replace("_", "")).strip()
    
    return normalized



def find_column_by_canon(df, target_name, aliases=None):
    """Return actual column name in df that matches target_name canonicalized.
    Uses exact match after canonicalization. If aliases provided, checks those too.
    
    Args:
        df: DataFrame to search
        target_name: Column name to find (canonical key like 'material' or actual name)
        aliases: Optional dict of canonical_key -> [alias1, alias2, ...] for flexible matching
        
    Returns:
        Actual column name in df, or None if not found
    """
    # Build reverse mapping: canon(header) -> actual header
    canon_to_header = {}
    for col in df.columns:
        ch = canon_header(col)
        if ch and ch not in canon_to_header:
            canon_to_header[ch] = col
    
    # First try: direct match with target_name
    target_canon = canon_header(target_name)
    if target_canon in canon_to_header:
        return canon_to_header[target_canon]
    
    # Second try: if aliases provided, check if target_name is a canonical key with aliases
    if aliases and target_canon in aliases:
        for alias in aliases[target_canon]:
            alias_canon = canon_header(alias)
            if alias_canon in canon_to_header:
                return canon_to_header[alias_canon]
    
    return None


def ensure_file_not_open(path: Path, label: str) :
    """
    Raise PermissionError if the file appears to be open/locked.
    """
    locked = file_in_use(path)
    if locked:
        raise PermissionError(
            f"The following {label} file looks open/locked:\n- " +
            f"{path.name}\n\nClose it and try again."
        )
    

# --- File locking check - mechanism 
def file_in_use(path) -> bool:
    """
    Returns True if the file appears to be locked/open by another process.
    Uses advisory locks where available; falls back to PermissionError.
    """
    if not os.path.isfile(path): 
        return True
    try:
        # Open read+binary to probe lock status
        with open(path, "rb+") as f:
            if os.name == "nt":
                import msvcrt
                try:
                    msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                    msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
                except OSError:
                    return True
        return False
    except PermissionError:
        return True
    except Exception:
        # If anything odd happens, be conservative
        return True



def compute_output_path(leading_path: Path) -> Path:
    """
    Compute output path for enriched report in the LEADING file's grandparent folder,
    in a subfolder "IB_MatchWise_Results". The filename includes a timestamp.
    """
    MAX_PATH = 260  # Windows path limit
    MIN_FILENAME_LENGTH = 20  # Ensure meaningful filename

    ts = datetime.now().strftime("%m-%d-%Y__%H-%M-%S")
    
    # Validate input path depth
    try:
        leading_parent = leading_path.parent.parent
    except (AttributeError, IndexError):
        # Fallback to parent if grandparent doesn't exist
        leading_parent = leading_path.parent

    # Build directory name with length checking
    res_folder_name = f"IB_MatchWise_Results__{ts}"
    base_filename = f"Results_Report__{ts}.xlsx"
    
    # Calculate available space for directory name
    base_path_len = len(str(leading_parent))
    remaining_space = MAX_PATH - base_path_len - len(base_filename) - 10  # buffer
    
    if len(res_folder_name) > remaining_space:
        # Shorten directory name intelligently
        res_folder_name = f"IB_Results__{ts}"
        if len(res_folder_name) > remaining_space:
            # Last resort: minimal directory name
            res_folder_name = f"IB__{ts}"

    out_dir = leading_parent / res_folder_name
    
    # Create directory with unique naming if conflict
    counter = 1
    original_dir = out_dir
    while out_dir.exists():
        out_dir = original_dir.parent / f"{original_dir.name}_{counter}"
        counter += 1
        if counter > 100:  # Prevent infinite loop
            break
    
    out_dir.mkdir(parents=True, exist_ok=True)
    
    # Build final path with validation
    final_path = out_dir / base_filename
    
    # Final length check
    if len(str(final_path)) > MAX_PATH:
        # Shorten filename while preserving extension
        available_for_name = MAX_PATH - len(str(out_dir)) - 6  # .xlsx + separator
        if available_for_name < MIN_FILENAME_LENGTH:
            # Directory path too long, use very short name
            final_filename = "report.xlsx"
        else:
            # Keep timestamp but shorten description
            short_name = f"Report_{ts}.xlsx"
            if len(short_name) > available_for_name:
                # Use minimal name with just date
                date_only = ts.split("__")[0]
                short_name = f"R_{date_only}.xlsx"
            final_filename = short_name
        
        final_path = out_dir / final_filename
    
    return final_path



def pick_sheet(path: Path, file_type: str) -> str:
    """
    For Excel files, return the sheet name to use. If "Sheet1" exists, use it. Otherwise, use the first sheet.
    For CSV files, return a default sheet name since they have no sheets.
    
    input:
        - path: Path to the file
        - file_type: Type of file (e.g., "cbom") for error messages
    output:
        - string: The name of the sheet to use for processing
    """
    # Convert to Path object if it's a string
    if isinstance(path, str):
        path = Path(path)
    
    ext = path.suffix.lower()
    sheetnames = []
    try:
        if ext == ".csv":
            # CSV files have no sheets
            sheetnames = [ "CSV_Sheet" ] 
        elif ext in [".xlsx", ".xlsm"]:
            # For .xlsx/.xlsm files, use openpyxl
            wb = load_workbook(path, read_only=True)
            sheetnames = [s.strip() for s in wb.sheetnames]
            wb.close()
        elif ext in [".xls"]:
            # For .xls files, use pandas to read sheet names
            xls = pd.ExcelFile(path)
            sheetnames = [s.strip() for s in xls.sheet_names]
        else:
            raise ValueError(f"Unsupported file format: {ext}. Only .xlsx, .xlsm, and .csv files are supported.")
    except Exception as e:
        raise ValueError(f"Error reading sheets from {path.name}: {str(e)}")
    
    if not sheetnames:
        raise ValueError(f"No sheets found in {path.name}")
    if (file_type == "cbom") :
        if len(sheetnames) > 1:
            config = load_config(config_path='config.json')
            target_sheet = config["cbom"]["target_sheet"].get("name", "C-BoM 830234")
            for name in sheetnames:
                if name.casefold() == target_sheet.casefold():
                    return name
            # If no match, show warning and raise error
            raise ValueError(f"Could not find  target_sheet in {path.name}")
    if file_type == "dictionary" :
        if len(sheetnames) > 1:
            config = load_config(config_path='config.json')
            target_sheet = config["dictionary"]["target_sheet"].get("name", "12NC_Mapping")
            for name in sheetnames:
                if name.casefold() == target_sheet.casefold():
                    return name
            # If no match, show warning and raise error
            raise ValueError(f"Could not find required sheet '{target_sheet}' in {path.name}")
    # Default: return first sheet
    return sheetnames[0]


def setup_logger(log_path: Path):
    """
    Create and configure a dedicated logger for this run.
    Ensures:
    - no duplicate handlers
    - correct log file every run
    - console + file output
    
    Returns configured logger instance.
    """
    logger = logging.getLogger("ib_matchwise")  # dedicated logger
    logger.setLevel(logging.INFO)
    
    # Prevent propagation to root logger to avoid duplicates
    logger.propagate = False

    # IMPORTANT: Remove old handlers if re-running
    if logger.hasHandlers():
        for handler in logger.handlers[:]:
            handler.close()
            logger.removeHandler(handler)

    # --- File Handler ---
    try:
        fh = logging.FileHandler(log_path, mode="w", encoding="utf-8")
        fh.setLevel(logging.INFO)
        formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    except Exception as e:
        # If file handler fails, log warning to console handler only
        logger.warning(f"Could not create log file: {e}")

    return logger



def canon_header(s: str) -> str:
    """
    Canonicalize a header for robust matching
    """
    if s is None:
        return ""
    s = s.casefold()
    s = s.strip(" ")
    s = s.strip("\n")
    s = s.strip("\t")
    s = s.strip("\r")
    s = s.replace("\n", " ").replace("\t", " ").replace("\r", " ")              
    return s


def is_blank(val, blank_tokens) -> bool:
    if pd.isna(val):
        return True
    s = str(val).strip()
    if s == "":
        return True
    return s.casefold() in {t.casefold() for t in blank_tokens}


def load_config(config_path='config.json'):
    """
    Load configuration from JSON file, or create default if not exists
    
    Input:
    - config_path: Path to config JSON file (default: 'config.json')
    
    Output:
    - Dictionary containing configuration settings
    """
   
    if os.path.exists(config_path):
        with open(config_path, 'r') as f:
            return json.load(f)
    else:
        # Create default config
        default_config = {
            'cbom_room_col_start': 'G',
            'cbom_room_num_row': 5,
            'cbom_room_desc_row': 4,
            'cbom_12nc_col': 'C',
            'cbom_12nc_desc_col': 'D',
            'cbom_12nc_row_start': 9
        }
        with open(config_path, 'w') as f:
            json.dump(default_config, f, indent=4)
        print(f"Created default config file: {config_path}")
        return default_config
