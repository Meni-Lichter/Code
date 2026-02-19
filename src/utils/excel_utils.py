"""Excel utility functions for sheet selection and column handling"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from .string_utils import canon_header


def pick_sheet(path: Path, file_type: str, config: dict) -> str:
    """
    For Excel files, return the sheet name to use. If "Sheet1" exists, use it. Otherwise, use the first sheet.
    For CSV files, return a default sheet name since they have no sheets.

    input:
        - path: Path to the file
        - file_type: Type of file (e.g., "cbom") for error messages
    output:
        - string: The name of the sheet to use for processing
    """
    # Convert to Path object if it's a string
    if isinstance(path, str):
        path = Path(path)

    ext = path.suffix.lower()
    sheetnames = []
    try:
        if ext == ".csv":
            # CSV files have no sheets
            sheetnames = ["CSV_Sheet"]
        elif ext in [".xlsx", ".xlsm"]:
            # For .xlsx/.xlsm files, use openpyxl
            wb = load_workbook(path, read_only=True)
            sheetnames = [str(s).strip() for s in wb.sheetnames]
            wb.close()
        elif ext in [".xls"]:
            # For .xls files, use pandas to read sheet names
            xls = pd.ExcelFile(path)
            sheetnames = [str(s).strip() for s in xls.sheet_names]
        else:
            raise ValueError(
                f"Unsupported file format: {ext}. Only .xlsx, .xlsm, and .csv files are supported."
            )
    except Exception as e:
        raise ValueError(f"Error reading sheets from {path.name}: {str(e)}")

    if not sheetnames:
        raise ValueError(f"No sheets found in {path.name}")
    if file_type == "cbom":
        if len(sheetnames) > 1:
            target_sheet = config["cbom"]["target_sheet"].get("name", "C-BoM 830234")
            for name in sheetnames:
                if name.casefold() == target_sheet.casefold():
                    return name
            # If no match, show warning and raise error
            raise ValueError(f"Could not find  target_sheet in {path.name}")
    if file_type == "fit_cvi":
        if len(sheetnames) > 1:
            target_sheet = config["fit_cvi"]["target_sheet"].get("name", "FIT_CVI")
            for name in sheetnames:
                if name.casefold() == target_sheet.casefold():
                    return name
            # If no match, show warning and raise error
            raise ValueError(f"Could not find required sheet '{target_sheet}' in {path.name}")
    if file_type == "ymbd":
        if len(sheetnames) > 1:
            target_sheet = config["ymbd"]["target_sheet"].get("name", "YMBD")
            for name in sheetnames:
                if name.casefold() == target_sheet.casefold():
                    return name
            # If no match, show warning and raise error
            raise ValueError(f"Could not find required sheet '{target_sheet}' in {path.name}")
    # Default: return first sheet
    return sheetnames[0]


def col_letter_to_index(col_letter):
    """Converts Excel column letter(s) to a zero-based index.
    Args:
        col_letter (str): Excel column letter(s) (e.g., 'A', 'B', ..., 'Z', 'AA', etc.)

    Returns:
        int: Zero-based column index
    """

    col_letter = col_letter.upper()
    index = 0
    for i, char in enumerate(reversed(col_letter)):
        index += (ord(char) - ord("A") + 1) * (26**i)
    return index - 1


def find_column_by_canon(df, target_name, aliases=None):
    """Return actual column name in df that matches target_name canonicalized.
    Uses exact match after canonicalization. If aliases provided, checks those too.

    Args:
        df: DataFrame to search
        target_name: Column name to find (canonical key like 'material' or actual name)
        aliases: Optional dict of canonical_key -> [alias1, alias2, ...] for flexible matching

    Returns:
        Actual column name in df, or None if not found
    """
    # Build reverse mapping: canon(header) -> actual header
    canon_to_header = {}
    for col in df.columns:
        ch = canon_header(col)
        if ch and ch not in canon_to_header:
            canon_to_header[ch] = col

    # First try: direct match with target_name
    target_canon = canon_header(target_name)
    if target_canon in canon_to_header:
        return canon_to_header[target_canon]

    # Second try: if aliases provided, check if target_name is a canonical key with aliases
    if aliases and target_canon in aliases:
        for alias in aliases[target_canon]:
            alias_canon = canon_header(alias)
            if alias_canon in canon_to_header:
                return canon_to_header[alias_canon]

    return None
